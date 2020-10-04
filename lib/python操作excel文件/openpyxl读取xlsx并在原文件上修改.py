from openpyxl import load_workbook
wb = load_workbook(filename=r'Book1.xlsx')
sheetnames = wb.sheetnames
print(sheetnames)
ws = wb[sheetnames[0]]
print(ws.cell(row=1, column=1).value)	#�����1,1��ʼ
ws.cell(row=1, column=2).value = 'wjh'
print(ws.cell(row=1, column=2).value)
wb.save(filename='Book1.xlsx')
