import json
from openpyxl import Workbook
wb = Workbook()
ws = wb.worksheets[0]
ws.title = "first sheet"


with open("es_goods.dump",'r', encoding='UTF-8') as file_object:
	contents = file_object.readlines()
	# print(type(contents), type(contents[0]))		# list, string
	
	# 提取出 keys，先写入excel表格的第一行
	# contents为list格式，其中每条元素为string格式，但其实string格式内是字典结构
	# 用json.loads()方法可以把字符串的""去掉，把引号内的字典结构提出来，存入 tmp
	tmp = json.loads(contents[0])
	# print(type(tmp))				# dict
	i = 1
	for k in tmp.keys():
		ws.cell(row=1, column=i, value=str(k))
		i += 1
		
	for index, element in enumerate(contents):
		a = json.loads(element)
		j = 1
		for v in a.values():
			ws.cell(row=index+2, column=j, value=str(v))
			j += 1
		
wb.save(filename='data.xlsx')

