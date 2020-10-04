#方法一：xls		坐标从0,0开始
import xlwt
xls = xlwt.Workbook()
sht1 = xls.add_sheet('Sheet1')
sht1.write(0, 0, '列1')
sht1.write(0, 1, '列2')
sht1.write(1, 0, 'a2')
sht1.write(1, 1, 'b2')
xls.save(r'C:\Task\test.xls')

#方法二：xlsx		坐标从1,1开始
from openpyxl import Workbook
wb = Workbook()
ws = wb.worksheets[0]
ws.title = "first sheet"
ws.cell(row=1, column=1, value='Hello World!')
wb.save(filename='test.xlsx')

# 移动至其他路径
fp = r"d:/temp/dataDir"
wb = xlwt.Workbook()
ws = wb.add_sheet("description")
ws.write(0, 0, 'test')
wb.save("template.xls") 
# wb.save(os.path.join(fp,"template.xls"))
shutil.move("template.xls",os.path.join(fp, "template.xls"))
