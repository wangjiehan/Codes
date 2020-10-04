# coding=utf-8
import os, sys

sys.path.append(os.path.join(sys.path[0], '../lib'))
import requests
from db_model import *
from func import *
from bs4 import BeautifulSoup
import re
from openpyxl import Workbook

_exp = ['(今|明|去|前|后).?年', '(本|下个)月', '(今|明|昨|前|后|近|当).?(日|天)', '(今|明|昨)(早|晚)', '很快(就|将)', '(即将|将要|将会|马上|很快|计划|已经|最终|结果|最后|与此同时)', '.*(日|号)(早|中|晚|上午|中午|下午|晚上|凌晨|傍晚|深夜|黎明)', '.*月.{0,2}(日|号)', '目前|之前|此前|以来|期间|随后|以后|后续|未来', '最新|最近', '(这|此).?次', '近(日|期|照)', '分钟', '正式', '召开|开幕|启动|首发|上演', '日讯', '本周', '周([1-5一二三四五六]|日)', '([1-9一二三四五六七八九十]).', '{0, 3}(日|天)']
_reg = [None] * len(_exp)
for i in range(len(_exp)):
    _reg[i] = re.compile(_exp[i])

def parse(cat, label_yes, label_no):
    conn2 = Conn.newConn()
    cur2 = conn2.cursor()
    check = {}
    articles = timeliness.where(status=2, category=cat).select()
    N_acc = 0           # 计算准确率的分母
    N_recall = 0        # 计算召回率的分母
    count_acc = 0       # 计算准确率的分子
    count_recall = 0    # 计算召回率的分子
    for a in articles:
        url = a.url
        id = int(a.id)
        c = str(a.content)
        tag = int(a.tag)
        flag = label_no         # flag 相当于正则预测结果
        strs = ''
        if c :
            N_acc += 1
            if tag == label_yes:
                N_recall += 1
            for reg in _reg:

                # 若正则中没有匹配到，则继续循环向下匹配，直到找到或循环结束。若一直没找到，flag 就保持 no 的标签
                match = reg.search(c)
                if match is None:
                    continue
                # 若正则中匹配到了，则 flag 更新为 yes 标签，不用再在正则中继续找下去，可直接跳出循环
                else:
                    flag = label_yes

                    highlight = reg.findall(c)
                    for elements in highlight:
                        for i in elements:
                            tmp = i
                            strs += tmp
                        strs += ' '
                strs += '、'
            strs.replace(' ', '')
            cur2.execute('UPDATE timeliness SET hightlights = %s WHERE url = %s', (strs[:-1], url))
            # 若 flag 和当前监督标记一致，则准确样本数 + 1
            if flag == tag:
                count_acc += 1
            # 若 flag 是 yes 标记（正例）且和当前监督标记一致，则召回样本数 + 1
            if flag == tag and flag == label_yes:
                count_recall += 1
            # 真实是正例，但却没预测出来的
            if tag == label_yes and flag != tag:
                check[id] = c
    acc = count_acc / N_acc
    recall = count_recall / N_recall
    conn2.commit()
    cur2.close()
    conn2.close()
    return acc, recall, check

def parse3(label_yes, label_no):
    conn2 = Conn.newConn()
    cur2 = conn2.cursor()
    check = {}
    articles = timeliness3.where().select()
    N_acc = 0           # 计算准确率的分母
    N_recall = 0        # 计算召回率的分母
    count_acc = 0       # 计算准确率的分子
    count_recall = 0    # 计算召回率的分子
    for a in articles:
        url = a.url
        c = str(a.content)
        tag = int(a.tag)
        flag = label_no
        strs = ''
        if c:       # 踢掉没爬出来content的样本，只计算有content的
            N_acc += 1
            if tag == label_yes:
                N_recall += 1
            for reg in _reg:
                # 若正则中没有匹配到，则继续循环向下匹配，直到找到或循环结束。若一直没找到，flag 就保持 no 的标签
                match = reg.search(c)
                if match is None:
                    continue
                # 若正则中匹配到了，则 flag 更新为 yes 标签，不用再在正则中继续找下去，可直接跳出循环
                else:
                    flag = label_yes
                    # 以下至SQL语句代码：将正则匹配出来的内容填入新的highlights列
                    # 将正则表达式中匹配到的正则项填入新的highlight列
                    reg = str(reg)
                    strs += reg[11:-1]
                    strs += '、'
                    '''
                    highlight = reg.findall(c)
                    for elements in highlight:
                        for i in elements:
                            tmp = i
                            strs += tmp
                        strs += ' '
                strs += '、'
            strs.replace(' ', '')
            '''
            cur2.execute('UPDATE timeliness3 SET hightlights = %s WHERE url = %s', (strs[:-1], url))
            # 若 flag 和当前监督标记一致，则准确样本数 + 1
            if flag == tag:
                count_acc += 1
            # 若 flag 是 yes 标记（正例）且和当前监督标记一致，则召回样本数 + 1
            if flag == tag and flag == label_yes:
                count_recall += 1
            # 真实是正例，但却没预测出来的
            if tag == label_yes and flag != tag:
                check[id] = c
    acc = count_acc / N_acc
    recall = count_recall / N_recall
    conn2.commit()
    cur2.close()
    conn2.close()
    return acc, recall, check

def outExcel(dict, filepath):
    wb = Workbook()
    ws = wb.worksheets[0]
    ws.title = "first sheet"
    i = 0
    for key, value in dict.items():
        i += 1
        ws.cell(row=i, column=1, value=key)
        ws.cell(row=i, column=2, value=value)
    wb.save(filename=filepath)

def refresh():
    articles = timeliness3.where().select()
    conn2 = Conn.newConn()
    cur2 = conn2.cursor()
    for t in articles:
        url = t.url
        cur2.execute('UPDATE timeliness3 SET url = %s WHERE url = %s', (url[:-1], url))
        # cur2.execute('UPDATE timeliness3 SET url = %s WHERE url = %s', (url+'l', url))
    conn2.commit()
    cur2.close()
    conn2.close()

def update_content():
    articles = timeliness3.where().select()
    conn2 = Conn.newConn()
    cur2 = conn2.cursor()
    for a in articles:
        url = a.url
        response = requests.get(url, headers={'User-Agent': 'Mozilla/5.0'}, stream=True)
        response.encoding = "utf-8"
        data = response.text
        soup = BeautifulSoup(data, features="lxml")
        content_div = soup.find_all("div", {"class": "content"})
        if content_div is not None and len(content_div) > 0:
            content = content_div[0].get_text()
            cur2.execute('UPDATE timeliness3 SET content = %s WHERE url = %s', (content, url))
            print('success at url', url)
    conn2.commit()
    cur2.close()
    conn2.close()

def simple_acc():
    articles = timeliness3.where(tag=1).select()
    conn2 = Conn.newConn()
    cur2 = conn2.cursor()
    N, count = 0, 0
    for a in articles:
        c = str(a.content)
        if c:
            N += 1
        url = a.url
        flag = False
        strs = ''
        for reg in _reg:
            match = reg.search(c)
            if match is None:
                continue
            # 正则匹配到，则预测有时效性
            else:
                flag = True

                highlight = reg.findall(c)
                for elements in highlight:
                    for i in elements:
                        tmp = i
                        strs += tmp
                    strs += ' '
            strs += '、'
        strs.replace(' ','')
        cur2.execute('UPDATE timeliness3 SET hightlights = %s WHERE url = %s', (strs[:-1], url))
        if flag is True:
            count += 1
    acc = count / N

    conn2.commit()
    cur2.close()
    conn2.close()

    return acc


if __name__ == '__main__':
    # （1）去掉url中.html后的空格（只执行一次就行）
    # refresh()

    # （2）爬取url上的content数据，填入content列
    # update_content()

    # （3）计算准确率和召回率，并将真正例中未能被预测出来的数据（1-recall部分）存入excel
    res = parse3(1, 0)
    print("Accurracy = %.4f , recall = %.4f." % (res[0], res[1]))
    # outExcel(res[2], r'C:\Users\24299\Desktop/part(1-recall).xlsx')

    # print("The accurracy is %.4f" % simple_acc())

    # cat_list = [[1, 9, 29, 45], [33, 52]]

    # res_1 = parse(1, 3, 0)
    # res_9 = parse(9, 3, 0)
    # res_29 = parse(29, 3, 0)
    # res_45 = parse(45, 3, 0)
    # res_33 = parse(33, 1, 3)
    # res_52 = parse(52, 1, 3)

    # print("Category 1: accurracy = %.4f , recall = %.4f." % (res_1[0], res_1[1]))
    # print("Category 9: accurracy = %.4f , recall = %.4f." % (res_9[0], res_9[1]))
    # print("Category 29: accurracy = %.4f , recall = %.4f." % (res_29[0], res_29[1]))
    # print("Category 45: accurracy = %.4f , recall = %.4f." % (res_45[0], res_45[1]))
    # print("Category 33: accurracy = %.4f , recall = %.4f." % (res_33[0], res_33[1]))
    # print("Category 52: accurracy = %.4f , recall = %.4f." % (res_52[0], res_52[1]))

    # outExcel(res_1[2], r'C:\Users\24299\Desktop/cat_1.xlsx')
    # outExcel(res_9[2], r'C:\Users\24299\Desktop/cat_9.xlsx')
    # outExcel(res_29[2], r'C:\Users\24299\Desktop/cat_29.xlsx')
    # outExcel(res_45[2], r'C:\Users\24299\Desktop/cat_45.xlsx')
    # outExcel(res_33[2], r'C:\Users\24299\Desktop/cat_33.xlsx')
    # outExcel(res_52[2], r'C:\Users\24299\Desktop/cat_52.xlsx')



